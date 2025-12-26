from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from django.db.models import Q, Count
from .models import (
    JenisKejahatan, NamaKejahatan, Kecamatan, Desa, Status,
    LaporanKejahatan, FotoLaporanKejahatan, PosKeamanan, FotoPosKeamanan,
    CCTV, FotoCCTV, KejadianLainnya, FotoKejadianLainnya, User, Area
)
from .serializers import (
    JenisKejahatanSerializer, NamaKejahatanSerializer, KecamatanSerializer,
    DesaSerializer, StatusSerializer, LaporanKejahatanSerializer,
    FotoLaporanKejahatanSerializer, PosKeamananSerializer, FotoPosKeamananSerializer,
    CCTVSerializer, FotoCCTVSerializer, KejadianLainnyaSerializer,
    FotoKejadianLainnyaSerializer, UserSerializer, UserCreateSerializer, UserUpdateSerializer,
    LoginSerializer, ChangePasswordSerializer
)

from django.contrib.gis.geos import GEOSGeometry
from django.contrib.gis.db.models.functions import AsGeoJSON
import json
from rest_framework.authtoken.models import Token
from django.contrib.auth import login, logout
from django.views.decorators.csrf import csrf_exempt
from django.db import models
from rest_framework import permissions
from datetime import datetime, timedelta 
from django.db.models.functions import TruncMonth
from django.utils import timezone
from django.http import HttpResponse, FileResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import zipfile
import io
import tempfile
import os
from django.contrib.gis.gdal import DataSource, OGRGeometry, SpatialReference
from django.contrib.gis.gdal import Driver
import shutil




class JenisKejahatanViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    """
    ViewSet untuk CRUD Jenis Kejahatan
    """
    queryset = JenisKejahatan.objects.all()
    serializer_class = JenisKejahatanSerializer

    def get_queryset(self):
        queryset = JenisKejahatan.objects.all()
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(nama_jenis_kejahatan__icontains=search) |
                Q(deskripsi__icontains=search)
            )
        return queryset


class NamaKejahatanViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    """
    ViewSet untuk CRUD Nama Kejahatan
    """
    queryset = NamaKejahatan.objects.select_related('jenis_kejahatan').all()
    serializer_class = NamaKejahatanSerializer

    def get_queryset(self):
        queryset = NamaKejahatan.objects.select_related('jenis_kejahatan').all()
        jenis_kejahatan_id = self.request.query_params.get('jenis_kejahatan_id', None)
        search = self.request.query_params.get('search', None)
        
        if jenis_kejahatan_id:
            queryset = queryset.filter(jenis_kejahatan_id=jenis_kejahatan_id)
        
        if search:
            queryset = queryset.filter(
                Q(nama__icontains=search) |
                Q(deskripsi__icontains=search)
            )
        return queryset


class KecamatanViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    """
    ViewSet untuk CRUD Kecamatan
    """
    queryset = Kecamatan.objects.all()
    serializer_class = KecamatanSerializer

    def get_queryset(self):
        queryset = Kecamatan.objects.all()
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(nama__icontains=search) |
                Q(deskripsi__icontains=search)
            )
        return queryset


class DesaViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    """
    ViewSet untuk CRUD Desa
    """
    queryset = Desa.objects.select_related('kecamatan').all()
    serializer_class = DesaSerializer

    def get_queryset(self):
        queryset = Desa.objects.select_related('kecamatan').all()
        kecamatan_id = self.request.query_params.get('kecamatan_id', None)
        search = self.request.query_params.get('search', None)
        
        if kecamatan_id:
            queryset = queryset.filter(kecamatan_id=kecamatan_id)
        
        if search:
            queryset = queryset.filter(
                Q(nama__icontains=search) |
                Q(deskripsi__icontains=search)
            )
        return queryset


class StatusViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    """
    ViewSet untuk CRUD Status
    """
    queryset = Status.objects.all()
    serializer_class = StatusSerializer


class LaporanKejahatanViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    """
    ViewSet untuk CRUD Laporan Kejahatan
    """
    queryset = LaporanKejahatan.objects.select_related(
        'jenis_kejahatan', 'nama_kejahatan', 'kecamatan', 'desa', 'status'
    ).prefetch_related('foto').all()
    serializer_class = LaporanKejahatanSerializer

    def get_queryset(self):
        queryset = LaporanKejahatan.objects.select_related(
            'jenis_kejahatan', 'nama_kejahatan', 'kecamatan', 'desa', 'status'
        ).prefetch_related('foto').all()
        
        # Filter parameters
        jenis_kejahatan_id = self.request.query_params.get('jenis_kejahatan_id', None)
        nama_kejahatan_id = self.request.query_params.get('nama_kejahatan_id', None)
        kecamatan_id = self.request.query_params.get('kecamatan_id', None)
        desa_id = self.request.query_params.get('desa_id', None)
        status_id = self.request.query_params.get('status_id', None)
        is_approval = self.request.query_params.get('is_approval', None)
        search = self.request.query_params.get('search', None)
        
        if jenis_kejahatan_id:
            queryset = queryset.filter(jenis_kejahatan_id=jenis_kejahatan_id)
        if nama_kejahatan_id:
            queryset = queryset.filter(nama_kejahatan_id=nama_kejahatan_id)
        if kecamatan_id:
            queryset = queryset.filter(kecamatan_id=kecamatan_id)
        if desa_id:
            queryset = queryset.filter(desa_id=desa_id)
        if status_id:
            queryset = queryset.filter(status_id=status_id)
        if is_approval is not None:
            queryset = queryset.filter(is_approval=is_approval.lower() == 'true')
        if search:
            queryset = queryset.filter(
                Q(nama_pelapor__icontains=search) |
                Q(alamat__icontains=search) |
                Q(deskripsi__icontains=search)
            )
        
        return queryset

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Endpoint untuk approve laporan"""
        laporan = self.get_object()
        laporan.is_approval = True
        laporan.save()
        serializer = self.get_serializer(laporan)
        return Response(serializer.data)
    

# Tambahkan action methods ini ke dalam class LaporanKejahatanViewSet

    @action(detail=False, methods=['get'], url_path='download-excel')
    def download_excel(self, request):
        """
        Download data laporan kejahatan dalam format Excel
        Endpoint: /api/laporan-kejahatan/download-excel/
        """
        try:
            # Get filtered queryset based on query parameters
            queryset = self.get_queryset()
            
            # Debug: Log jumlah data
            print(f"Download Excel - Total records: {queryset.count()}")
            
            # Jika tidak ada data, return error response
            if not queryset.exists():
                return Response(
                    {'error': 'Tidak ada data untuk diunduh dengan filter yang dipilih'}, 
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Data Kriminalitas"
            
            # Define styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = [
                'ID', 'Nama Pelapor', 'Jenis Kejahatan', 'Nama Kejahatan',
                'Tanggal Kejadian', 'Waktu Kejadian', 'Kecamatan', 'Desa',
                'Alamat', 'Deskripsi', 'Status', 'Latitude', 'Longitude',
                'Approval', 'Tanggal Dibuat'
            ]
            
            # Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Write data
            row_num = 2
            for laporan in queryset:
                # Get coordinates
                latitude = laporan.lokasi.y if laporan.lokasi else ''
                longitude = laporan.lokasi.x if laporan.lokasi else ''
                
                row_data = [
                    laporan.id,
                    laporan.nama_pelapor,
                    laporan.jenis_kejahatan.nama_jenis_kejahatan,
                    laporan.nama_kejahatan.nama,
                    laporan.tanggal_kejadian.strftime('%d-%m-%Y'),
                    laporan.waktu_kejadian.strftime('%H:%M'),
                    laporan.kecamatan.nama,
                    laporan.desa.nama,
                    laporan.alamat,
                    laporan.deskripsi,
                    laporan.status.nama,
                    latitude,
                    longitude,
                    'Ya' if laporan.is_approval else 'Tidak',
                    laporan.created_at.strftime('%d-%m-%Y %H:%M')
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                
                row_num += 1
            
            # Adjust column widths
            column_widths = [8, 25, 20, 25, 15, 12, 15, 15, 30, 40, 15, 12, 12, 10, 18]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # Freeze first row
            ws.freeze_panes = 'A2'
            
            # Save to response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=data_kriminalitas_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(response)
            
            print(f"Download Excel - Success! Total rows: {row_num - 2}")
            return response
            
        except Exception as e:
            print(f"Download Excel - Error: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Gagal mengunduh Excel: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='download-geojson')
    def download_geojson(self, request):
        """
        Download data laporan kejahatan dalam format GeoJSON
        Endpoint: /api/laporan-kejahatan/download-geojson/
        """
        try:
            # Get filtered queryset based on query parameters
            queryset = self.get_queryset()
            
            # Debug: Log jumlah data
            print(f"Download GeoJSON - Total records: {queryset.count()}")
            
            # Build GeoJSON structure
            features = []
            for laporan in queryset:
                if laporan.lokasi:  # Only include records with location
                    feature = {
                        "type": "Feature",
                        "geometry": {
                            "type": "Point",
                            "coordinates": [laporan.lokasi.x, laporan.lokasi.y]
                        },
                        "properties": {
                            "id": laporan.id,
                            "nama_pelapor": laporan.nama_pelapor,
                            "jenis_kejahatan": laporan.jenis_kejahatan.nama_jenis_kejahatan,
                            "nama_kejahatan": laporan.nama_kejahatan.nama,
                            "tanggal_kejadian": laporan.tanggal_kejadian.strftime('%Y-%m-%d'),
                            "waktu_kejadian": laporan.waktu_kejadian.strftime('%H:%M:%S'),
                            "kecamatan": laporan.kecamatan.nama,
                            "desa": laporan.desa.nama,
                            "alamat": laporan.alamat,
                            "deskripsi": laporan.deskripsi,
                            "status": laporan.status.nama,
                            "is_approval": laporan.is_approval,
                            "created_at": laporan.created_at.strftime('%Y-%m-%d %H:%M:%S')
                        }
                    }
                    features.append(feature)
            
            print(f"Download GeoJSON - Total features with location: {len(features)}")
            
            geojson = {
                "type": "FeatureCollection",
                "name": "Data Kriminalitas",
                "crs": {
                    "type": "name",
                    "properties": {
                        "name": "urn:ogc:def:crs:OGC:1.3:CRS84"
                    }
                },
                "features": features
            }
            
            # Create response
            response = HttpResponse(
                json.dumps(geojson, indent=2, ensure_ascii=False),
                content_type='application/geo+json'
            )
            response['Content-Disposition'] = f'attachment; filename=data_kriminalitas_{timezone.now().strftime("%Y%m%d_%H%M%S")}.geojson'
            
            print(f"Download GeoJSON - Success!")
            return response
            
        except Exception as e:
            print(f"Download GeoJSON - Error: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Gagal mengunduh GeoJSON: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='download-shapefile')
    def download_shapefile(self, request):
        """
        Download data laporan kejahatan dalam format Shapefile (zipped)
        Endpoint: /api/laporan-kejahatan/download-shapefile/
        """
        try:
            # Get filtered queryset based on query parameters
            queryset = self.get_queryset().filter(lokasi__isnull=False)
            
            print(f"Download Shapefile - Total records with location: {queryset.count()}")
            
            if not queryset.exists():
                return Response(
                    {'error': 'Tidak ada data dengan lokasi untuk diunduh'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Create temporary directory for shapefile
            temp_dir = tempfile.mkdtemp()
            shapefile_name = f'data_kriminalitas_{timezone.now().strftime("%Y%m%d_%H%M%S")}'
            shapefile_path = os.path.join(temp_dir, shapefile_name)
            
            try:
                # Create shapefile using GDAL
                driver = Driver('ESRI Shapefile')
                ds = driver.CreateDataSource(shapefile_path + '.shp')
                
                # Create spatial reference (WGS84)
                srs = SpatialReference(4326)
                
                # Create layer
                layer = ds.CreateLayer('kriminalitas', srs, 1)  # 1 = wkbPoint
                
                # Define fields
                from django.contrib.gis.gdal import OGRFieldDefn
                fields = [
                    ('id', 0),  # 0 = OFTInteger
                    ('nama_pelap', 2),  # 2 = OFTString
                    ('jenis_krj', 2),
                    ('nama_krj', 2),
                    ('tgl_kjdn', 2),
                    ('waktu_kjdn', 2),
                    ('kecamatan', 2),
                    ('desa', 2),
                    ('alamat', 2),
                    ('deskripsi', 2),
                    ('status', 2),
                    ('approval', 2),
                    ('created_at', 2),
                ]
                
                for field_name, field_type in fields:
                    field_defn = OGRFieldDefn(field_name, field_type)
                    if field_type == 2:  # String
                        field_defn.width = 254
                    layer.create_field(field_defn)
                
                # Add features
                feature_count = 0
                for laporan in queryset:
                    # Create feature
                    feature = layer.feature_defn
                    feat = feature.create()
                    
                    # Set geometry
                    geom = OGRGeometry('POINT({} {})'.format(
                        laporan.lokasi.x, 
                        laporan.lokasi.y
                    ))
                    feat.geom = geom
                    
                    # Set attributes
                    feat['id'] = laporan.id
                    feat['nama_pelap'] = laporan.nama_pelapor[:254]
                    feat['jenis_krj'] = laporan.jenis_kejahatan.nama_jenis_kejahatan[:254]
                    feat['nama_krj'] = laporan.nama_kejahatan.nama[:254]
                    feat['tgl_kjdn'] = laporan.tanggal_kejadian.strftime('%Y-%m-%d')
                    feat['waktu_kjdn'] = laporan.waktu_kejadian.strftime('%H:%M:%S')
                    feat['kecamatan'] = laporan.kecamatan.nama[:254]
                    feat['desa'] = laporan.desa.nama[:254]
                    feat['alamat'] = laporan.alamat[:254]
                    feat['deskripsi'] = laporan.deskripsi[:254]
                    feat['status'] = laporan.status.nama[:254]
                    feat['approval'] = 'Ya' if laporan.is_approval else 'Tidak'
                    feat['created_at'] = laporan.created_at.strftime('%Y-%m-%d')
                    
                    # Add feature to layer
                    layer.add(feat)
                    feature_count += 1
                
                print(f"Download Shapefile - Added {feature_count} features")
                
                # Close datasource to flush to disk
                ds = None
                
                # Create zip file
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    # Add all shapefile components
                    for ext in ['.shp', '.shx', '.dbf', '.prj', '.cpg']:
                        file_path = shapefile_path + ext
                        if os.path.exists(file_path):
                            zip_file.write(file_path, shapefile_name + ext)
                            print(f"Added to zip: {shapefile_name}{ext}")
                
                # Clean up temp directory
                shutil.rmtree(temp_dir)
                
                # Return zip file
                zip_buffer.seek(0)
                response = HttpResponse(zip_buffer.read(), content_type='application/zip')
                response['Content-Disposition'] = f'attachment; filename={shapefile_name}.zip'
                
                print(f"Download Shapefile - Success!")
                return response
                
            except Exception as e:
                # Clean up on error
                if os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir)
                raise e
            
        except Exception as e:
            print(f"Download Shapefile - Error: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response(
                {'error': f'Gagal mengunduh Shapefile: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



class FotoLaporanKejahatanViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    """
    ViewSet untuk CRUD Foto Laporan Kejahatan
    """
    queryset = FotoLaporanKejahatan.objects.all()
    serializer_class = FotoLaporanKejahatanSerializer

    def get_queryset(self):
        queryset = FotoLaporanKejahatan.objects.all()
        laporan_kejahatan_id = self.request.query_params.get('laporan_kejahatan_id', None)
        
        if laporan_kejahatan_id:
            queryset = queryset.filter(laporan_kejahatan_id=laporan_kejahatan_id)
        
        return queryset


class PosKeamananViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    """
    ViewSet untuk CRUD Pos Keamanan
    """
    queryset = PosKeamanan.objects.select_related('desa__kecamatan').prefetch_related('foto').all()
    serializer_class = PosKeamananSerializer

    def get_queryset(self):
        queryset = PosKeamanan.objects.select_related('desa__kecamatan').prefetch_related('foto').all()
        
        desa_id = self.request.query_params.get('desa_id', None)
        search = self.request.query_params.get('search', None)
        
        if desa_id:
            queryset = queryset.filter(desa_id=desa_id)
        if search:
            queryset = queryset.filter(
                Q(nama__icontains=search) |
                Q(alamat__icontains=search) |
                Q(keterangan__icontains=search)
            )
        
        return queryset


class FotoPosKeamananViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    """
    ViewSet untuk CRUD Foto Pos Keamanan
    """
    queryset = FotoPosKeamanan.objects.all()
    serializer_class = FotoPosKeamananSerializer

    def get_queryset(self):
        queryset = FotoPosKeamanan.objects.all()
        pos_keamanan_id = self.request.query_params.get('pos_keamanan_id', None)
        
        if pos_keamanan_id:
            queryset = queryset.filter(pos_keamanan_id=pos_keamanan_id)
        
        return queryset


class CCTVViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    """
    ViewSet untuk CRUD CCTV
    """
    queryset = CCTV.objects.select_related('kecamatan', 'desa').prefetch_related('foto').all()
    serializer_class = CCTVSerializer

    def get_queryset(self):
        queryset = CCTV.objects.select_related('kecamatan', 'desa').prefetch_related('foto').all()
        
        kecamatan_id = self.request.query_params.get('kecamatan_id', None)
        desa_id = self.request.query_params.get('desa_id', None)
        search = self.request.query_params.get('search', None)
        
        if kecamatan_id:
            queryset = queryset.filter(kecamatan_id=kecamatan_id)
        if desa_id:
            queryset = queryset.filter(desa_id=desa_id)
        if search:
            queryset = queryset.filter(
                Q(nama_lokasi__icontains=search) |
                Q(deskripsi__icontains=search)
            )
        
        return queryset


class FotoCCTVViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    """
    ViewSet untuk CRUD Foto CCTV
    """
    queryset = FotoCCTV.objects.all()
    serializer_class = FotoCCTVSerializer

    def get_queryset(self):
        queryset = FotoCCTV.objects.all()
        cctv_id = self.request.query_params.get('cctv_id', None)
        
        if cctv_id:
            queryset = queryset.filter(cctv_id=cctv_id)
        
        return queryset


class KejadianLainnyaViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    """
    ViewSet untuk CRUD Kejadian Lainnya
    """
    queryset = KejadianLainnya.objects.select_related(
        'kecamatan', 'desa', 'status'
    ).prefetch_related('foto').all()
    serializer_class = KejadianLainnyaSerializer

    def get_queryset(self):
        queryset = KejadianLainnya.objects.select_related(
            'kecamatan', 'desa', 'status'
        ).prefetch_related('foto').all()
        
        kecamatan_id = self.request.query_params.get('kecamatan_id', None)
        desa_id = self.request.query_params.get('desa_id', None)
        status_id = self.request.query_params.get('status_id', None)
        is_approval = self.request.query_params.get('is_approval', None)
        search = self.request.query_params.get('search', None)
        
        if kecamatan_id:
            queryset = queryset.filter(kecamatan_id=kecamatan_id)
        if desa_id:
            queryset = queryset.filter(desa_id=desa_id)
        if status_id:
            queryset = queryset.filter(status_id=status_id)
        if is_approval is not None:
            queryset = queryset.filter(is_approval=is_approval.lower() == 'true')
        if search:
            queryset = queryset.filter(
                Q(nama_pelapor__icontains=search) |
                Q(nama_kejadian__icontains=search) |
                Q(deskripsi_kejadian__icontains=search)
            )
        
        return queryset

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Endpoint untuk approve kejadian"""
        kejadian = self.get_object()
        kejadian.is_approval = True
        kejadian.save()
        serializer = self.get_serializer(kejadian)
        return Response(serializer.data)


class FotoKejadianLainnyaViewSet(viewsets.ModelViewSet):
    permission_classes = [permissions.AllowAny]
    """
    ViewSet untuk CRUD Foto Kejadian Lainnya
    """
    queryset = FotoKejadianLainnya.objects.all()
    serializer_class = FotoKejadianLainnyaSerializer

    def get_queryset(self):
        queryset = FotoKejadianLainnya.objects.all()
        kejadian_lainnya_id = self.request.query_params.get('kejadian_lainnya_id', None)
        
        if kejadian_lainnya_id:
            queryset = queryset.filter(kejadian_lainnya_id=kejadian_lainnya_id)
        
        return queryset


@api_view(['GET'])
def statistik_dashboard(request):
    """
    API endpoint untuk mendapatkan statistik dashboard:
    - Jumlah Laporan Kejahatan yang sudah diapprove (is_approval = True)
    - Jumlah Desa
    - Jumlah Pos Keamanan
    - Jumlah CCTV
    """
    try:
        # 1. Jumlah Laporan Kejahatan yang diapprove
        jumlah_laporan_kejahatan_approved = LaporanKejahatan.objects.filter(
            is_approval=True
        ).count()
        
        # 2. Jumlah Desa
        jumlah_desa = Desa.objects.count()
        
        # 3. Jumlah Pos Keamanan
        jumlah_pos_keamanan = PosKeamanan.objects.count()
        
        # 4. Jumlah CCTV
        jumlah_cctv = CCTV.objects.count()
        
        # Menyusun response data
        data = {
            'success': True,
            'message': 'Statistik berhasil diambil',
            'data': {
                'jumlah_laporan_kejahatan_approved': jumlah_laporan_kejahatan_approved,
                'jumlah_desa': jumlah_desa,
                'jumlah_pos_keamanan': jumlah_pos_keamanan,
                'jumlah_cctv': jumlah_cctv
            }
        }
        
        return Response(data, status=status.HTTP_200_OK)
    
    except Exception as e:
        return Response({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}',
            'data': None
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


@api_view(['GET'])
def map_data(request):
    """
    API endpoint untuk mendapatkan semua data peta:
    - Areas (polygon dengan warna berbeda)
    - Laporan Kejahatan (markers)
    - Pos Keamanan (markers)
    - CCTV (markers)
    """
    try:
        # 1. Ambil data Areas dengan geometry
        from .models import Area
        areas = Area.objects.select_related('kecamatan', 'desa').all()
        
        # Generate warna untuk setiap area
        colors = [
            '#ef4444',  # red
            '#f59e0b',  # amber
            '#10b981',  # emerald
            '#3b82f6',  # blue
            '#8b5cf6',  # violet
            '#ec4899',  # pink
            '#14b8a6',  # teal
            '#f97316',  # orange
        ]
        
        # Konversi areas ke GeoJSON
        area_features = []
        for idx, area in enumerate(areas):
            if area.geom:
                feature = {
                    'type': 'Feature',
                    'id': area.id,
                    'properties': {
                        'id': area.id,
                        'name': area.wadmkd or area.desa.nama,
                        'kecamatan': area.kecamatan.nama,
                        'desa': area.desa.nama,
                        'luas': float(area.luas) if area.luas else 0,
                        'color': colors[idx % len(colors)],
                    },
                    'geometry': json.loads(area.geom.geojson)
                }
                area_features.append(feature)
        
        areas_geojson = {
            'type': 'FeatureCollection',
            'features': area_features
        }
        
        # 2. Ambil Laporan Kejahatan yang sudah di-approve
        from .models import LaporanKejahatan
        laporan_kejahatan = LaporanKejahatan.objects.filter(
            is_approval=True,
            lokasi__isnull=False
        ).select_related(
            'jenis_kejahatan', 'nama_kejahatan', 'kecamatan', 'desa', 'status'
        ).prefetch_related('foto')
        
        crime_markers = []
        for laporan in laporan_kejahatan:
            if laporan.lokasi:
                crime_markers.append({
                    'id': laporan.id,
                    'type': 'crime',
                    'name': laporan.nama_kejahatan.nama,
                    'jenis': laporan.jenis_kejahatan.nama_jenis_kejahatan,
                    'latitude': laporan.lokasi.y,
                    'longitude': laporan.lokasi.x,
                    'address': laporan.alamat,
                    'description': laporan.deskripsi,
                    'date': str(laporan.tanggal_kejadian),
                    'time': str(laporan.waktu_kejadian),
                    'status': laporan.status.nama,
                    'kecamatan': laporan.kecamatan.nama,
                    'desa': laporan.desa.nama,
                    'pelapor': laporan.nama_pelapor,
                    'photos': [foto.file_path.url for foto in laporan.foto.all()]
                })
        
        # 3. Ambil Pos Keamanan
        from .models import PosKeamanan
        pos_keamanan = PosKeamanan.objects.filter(
            lokasi__isnull=False
        ).select_related('desa__kecamatan').prefetch_related('foto')
        
        security_markers = []
        for pos in pos_keamanan:
            if pos.lokasi:
                security_markers.append({
                    'id': pos.id,
                    'type': 'security_post',
                    'name': pos.nama,
                    'latitude': pos.lokasi.y,
                    'longitude': pos.lokasi.x,
                    'address': pos.alamat,
                    'description': pos.keterangan or '',
                    'kecamatan': pos.desa.kecamatan.nama,
                    'desa': pos.desa.nama,
                    'photos': [foto.file_path.url for foto in pos.foto.all()]
                })
        
        # 4. Ambil CCTV
        from .models import CCTV
        cctv_list = CCTV.objects.filter(
            lokasi__isnull=False
        ).select_related('kecamatan', 'desa').prefetch_related('foto')
        
        cctv_markers = []
        for cctv in cctv_list:
            if cctv.lokasi:
                cctv_markers.append({
                    'id': cctv.id,
                    'type': 'cctv',
                    'name': cctv.nama_lokasi,
                    'latitude': cctv.lokasi.y,
                    'longitude': cctv.lokasi.x,
                    'description': cctv.deskripsi or '',
                    'url_cctv': cctv.url_cctv or '',
                    'kecamatan': cctv.kecamatan.nama,
                    'desa': cctv.desa.nama,
                    'photos': [foto.file_path.url for foto in cctv.foto.all()]
                })
        
        # Response data
        data = {
            'success': True,
            'message': 'Data peta berhasil diambil',
            'data': {
                'areas': areas_geojson,
                'crime_reports': crime_markers,
                'security_posts': security_markers,
                'cctvs': cctv_markers,
                'summary': {
                    'total_areas': len(area_features),
                    'total_crime_reports': len(crime_markers),
                    'total_security_posts': len(security_markers),
                    'total_cctvs': len(cctv_markers),
                }
            }
        }
        
        return Response(data, status=status.HTTP_200_OK)
    
    except Exception as e:
        return Response({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}',
            'data': None
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def area_statistics(request, area_id=None):
    """
    API endpoint untuk mendapatkan statistik area tertentu
    """
    try:
        from .models import Area, LaporanKejahatan, PosKeamanan, CCTV
        from django.contrib.gis.geos import GEOSGeometry
        
        if area_id:
            # Statistik untuk area spesifik
            area = Area.objects.get(id=area_id)
            
            # Hitung laporan kejahatan dalam area
            # Menggunakan spatial query untuk cek point dalam polygon
            crime_in_area = LaporanKejahatan.objects.filter(
                is_approval=True,
                lokasi__isnull=False,
                lokasi__within=area.geom
            ).select_related('jenis_kejahatan', 'nama_kejahatan', 'status')
            
            # Hitung pos keamanan dalam area
            security_in_area = PosKeamanan.objects.filter(
                lokasi__isnull=False,
                lokasi__within=area.geom
            )
            
            # Hitung CCTV dalam area
            cctv_in_area = CCTV.objects.filter(
                lokasi__isnull=False,
                lokasi__within=area.geom
            )
            
            # Statistik berdasarkan jenis kejahatan
            crime_by_type = crime_in_area.values(
                'jenis_kejahatan__nama_jenis_kejahatan'
            ).annotate(
                count=Count('id')
            ).order_by('-count')
            
            # Statistik berdasarkan status
            crime_by_status = crime_in_area.values(
                'status__nama'
            ).annotate(
                count=Count('id')
            )
            
            # Format data untuk chart
            by_type = [
                {
                    'name': item['jenis_kejahatan__nama_jenis_kejahatan'],
                    'value': item['count']
                }
                for item in crime_by_type
            ]
            
            solved = crime_by_status.filter(status__nama='Selesai').first()
            pending = crime_by_status.exclude(status__nama='Selesai').aggregate(
                total=Count('id')
            )
            
            # Recent cases
            recent_cases = []
            for crime in crime_in_area.order_by('-created_at')[:10]:
                recent_cases.append({
                    'id': crime.id,
                    'type': crime.nama_kejahatan.nama,
                    'location': crime.alamat,
                    'date': str(crime.tanggal_kejadian),
                    'status': crime.status.nama
                })
            
            # Monthly trend (simulasi - sesuaikan dengan kebutuhan)
            from django.db.models.functions import TruncMonth
            monthly_data = crime_in_area.annotate(
                month=TruncMonth('tanggal_kejadian')
            ).values('month').annotate(
                cases=Count('id')
            ).order_by('month')
            
            monthly_trend = [
                {
                    'month': item['month'].strftime('%b'),
                    'cases': item['cases']
                }
                for item in monthly_data
            ]
            
            statistics = {
                'area_info': {
                    'id': area.id,
                    'name': area.wadmkd or area.desa.nama,
                    'kecamatan': area.kecamatan.nama,
                    'desa': area.desa.nama,
                    'luas': float(area.luas) if area.luas else 0,
                },
                'crime_stats': {
                    'totalCases': crime_in_area.count(),
                    'solved': solved['count'] if solved else 0,
                    'pending': pending['total'] or 0,
                    'byType': by_type,
                    'recentCases': recent_cases,
                    'monthlyTrend': monthly_trend
                },
                'infrastructure': {
                    'security_posts': security_in_area.count(),
                    'cctvs': cctv_in_area.count()
                }
            }
            
            return Response({
                'success': True,
                'message': 'Statistik area berhasil diambil',
                'data': statistics
            }, status=status.HTTP_200_OK)
        
        else:
            # Statistik semua area
            areas = Area.objects.all()
            area_stats = []
            
            for area in areas:
                crime_count = LaporanKejahatan.objects.filter(
                    is_approval=True,
                    lokasi__isnull=False,
                    lokasi__within=area.geom
                ).count()
                
                area_stats.append({
                    'id': area.id,
                    'name': area.wadmkd or area.desa.nama,
                    'crime_count': crime_count,
                })
            
            return Response({
                'success': True,
                'message': 'Statistik semua area berhasil diambil',
                'data': area_stats
            }, status=status.HTTP_200_OK)
    
    except Area.DoesNotExist:
        return Response({
            'success': False,
            'message': 'Area tidak ditemukan',
            'data': None
        }, status=status.HTTP_404_NOT_FOUND)
    
    except Exception as e:
        return Response({
            'success': False,
            'message': f'Terjadi kesalahan: {str(e)}',
            'data': None
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class AreaViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet untuk CRUD Area (Read-only karena managed=False)
    """
    from .models import Area
    from .serializers import AreaSerializer
    
    queryset = Area.objects.select_related('kecamatan', 'desa').all()
    serializer_class = AreaSerializer
    
    @action(detail=True, methods=['get'])
    def statistics(self, request, pk=None):
        """Endpoint untuk mendapatkan statistik area"""
        return area_statistics(request, area_id=pk)
    



class IsAdminOrReadOnly(permissions.BasePermission):
    """
    Permission kustom: Admin bisa semua, user lain hanya read
    """
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return request.user and request.user.is_authenticated
        return request.user and request.user.is_authenticated and request.user.jabatan == 'admin'


class UserViewSet(viewsets.ModelViewSet):
    """
    ViewSet untuk CRUD User
    """
    queryset = User.objects.all()
    permission_classes = [permissions.IsAuthenticated, IsAdminOrReadOnly]
    
    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return UserUpdateSerializer
        return UserSerializer
    
    def get_queryset(self):
        queryset = User.objects.all()
        
        # Filter berdasarkan status
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        # Filter berdasarkan jabatan
        jabatan = self.request.query_params.get('jabatan', None)
        if jabatan:
            queryset = queryset.filter(jabatan=jabatan)
        
        # Search
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                models.Q(username__icontains=search) |
                models.Q(name__icontains=search) |
                models.Q(email__icontains=search)
            )
        
        return queryset.order_by('-created_at')
    
    @action(detail=True, methods=['post'])
    def change_password(self, request, pk=None):
        """Endpoint untuk ganti password user tertentu (admin only)"""
        user = self.get_object()
        
        if request.user.jabatan != 'admin' and request.user.id != user.id:
            return Response(
                {'error': 'Anda tidak memiliki akses'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        serializer = ChangePasswordSerializer(
            data=request.data,
            context={'request': request}
        )
        
        if serializer.is_valid():
            serializer.save()
            return Response({'message': 'Password berhasil diubah'})
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['post'])
    def toggle_status(self, request, pk=None):
        """Toggle status aktif/nonaktif user"""
        user = self.get_object()
        user.is_active = not user.is_active
        user.save()
        
        return Response({
            'message': f"User {'diaktifkan' if user.is_active else 'dinonaktifkan'}",
            'is_active': user.is_active
        })


@csrf_exempt  # ← TAMBAHKAN INI untuk bypass CSRF check pada login
@api_view(['POST'])
@permission_classes([permissions.AllowAny])
def login_view(request):
    """
    Login endpoint
    POST /api/auth/login/
    Body: { "username": "admin", "password": "admin123" }
    
    Returns:
    {
        "token": "9944b09199c62bcf9418ad846dd0e4bbdfc6ee4b",
        "user": {
            "id": 1,
            "username": "admin",
            "email": "admin@kriminalitas.com",
            "name": "Administrator",
            "jabatan": "admin"
        },
        "message": "Login berhasil"
    }
    """
    try:
        # Log request untuk debugging
        print(f"Login attempt - Username: {request.data.get('username')}")
        
        serializer = LoginSerializer(data=request.data)
        
        if serializer.is_valid():
            user = serializer.validated_data['user']
            
            # Hapus token lama jika ada
            Token.objects.filter(user=user).delete()
            
            # Buat token baru
            token = Token.objects.create(user=user)
            
            # OPTIONAL: Login user untuk session (tidak diperlukan untuk token auth)
            # login(request, user)
            
            print(f"Login successful - User: {user.username}")
            
            return Response({
                'token': token.key,
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'name': user.name,
                    'jabatan': user.jabatan,
                },
                'message': 'Login berhasil'
            }, status=status.HTTP_200_OK)
        
        # Invalid credentials
        print(f"Login failed - Errors: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
    except Exception as e:
        # Catch any unexpected errors
        print(f"Login error: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return Response({
            'error': 'Terjadi kesalahan server',
            'detail': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@csrf_exempt  # ← TAMBAHKAN INI juga untuk logout
@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def logout_view(request):
    """
    Logout endpoint
    POST /api/auth/logout/
    Headers: Authorization: Token <token>
    """
    try:
        # Hapus token
        request.user.auth_token.delete()
        logout(request)
        return Response({'message': 'Logout berhasil'})
    except Exception as e:
        print(f"Logout error: {str(e)}")
        return Response(
            {'error': 'Logout gagal', 'detail': str(e)}, 
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def current_user(request):
    """
    Get current logged in user
    GET /api/auth/me/
    Headers: Authorization: Token <token>
    """
    serializer = UserSerializer(request.user)
    return Response(serializer.data)


@api_view(['PUT'])
@permission_classes([permissions.IsAuthenticated])
def update_profile(request):
    """
    Update current user profile
    PUT /api/auth/profile/
    Headers: Authorization: Token <token>
    """
    serializer = UserUpdateSerializer(
        request.user, 
        data=request.data, 
        partial=True
    )
    
    if serializer.is_valid():
        serializer.save()
        return Response({
            'user': UserSerializer(request.user).data,
            'message': 'Profile berhasil diupdate'
        })
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def change_password_view(request):
    """
    Change password for current user
    POST /api/auth/change-password/
    Headers: Authorization: Token <token>
    Body: {
        "old_password": "old",
        "new_password": "new",
        "new_password_confirm": "new"
    }
    """
    serializer = ChangePasswordSerializer(
        data=request.data,
        context={'request': request}
    )
    
    if serializer.is_valid():
        serializer.save()
        return Response({'message': 'Password berhasil diubah'})
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def get_area_classifications(request):
    """
    Endpoint untuk mendapatkan klasifikasi semua area berdasarkan jumlah kasus
    GET /api/map/area-classifications/
    
    Returns:
    {
        "success": true,
        "data": {
            "areas": [
                {
                    "id": 1,
                    "name": "Desa A",
                    "total_cases": 7,
                    "classification": "TINGGI",
                    "color": "#ef4444"
                },
                ...
            ],
            "thresholds": {
                "low_max": 2,
                "medium_max": 5,
                "high_min": 6
            }
        }
    }
    """
    try:
        # Ambil semua area dengan desa
        areas = Area.objects.select_related('desa', 'kecamatan').all()
        
        # Hitung jumlah kasus per area
        area_data = []
        case_counts = []
        
        for area in areas:
            desa_id = area.desa_id
            total_cases = LaporanKejahatan.objects.filter(desa_id=desa_id).count()
            
            area_data.append({
                'id': area.id,
                'name': area.desa.nama if area.desa else area.wadmkd,
                'kecamatan': area.kecamatan.nama if area.kecamatan else area.wadmkc,
                'total_cases': total_cases,
                'desa_id': desa_id
            })
            case_counts.append(total_cases)
        
        # Jika tidak ada data
        if not case_counts:
            return Response({
                'success': True,
                'data': {
                    'areas': [],
                    'thresholds': {'low_max': 0, 'medium_max': 0, 'high_min': 0}
                }
            })
        
        # Hitung threshold menggunakan percentile (33% dan 67%)
        import numpy as np
        case_counts_sorted = sorted(case_counts)
        
        # Percentile 33% dan 67%
        low_threshold = np.percentile(case_counts_sorted, 33)
        medium_threshold = np.percentile(case_counts_sorted, 67)
        
        # Definisikan warna
        COLOR_LOW = "#10b981"     # Hijau (green-500)
        COLOR_MEDIUM = "#f59e0b"  # Kuning (amber-500)
        COLOR_HIGH = "#ef4444"    # Merah (red-500)
        
        # Klasifikasikan setiap area
        for item in area_data:
            total = item['total_cases']
            
            if total <= low_threshold:
                item['classification'] = 'RENDAH'
                item['color'] = COLOR_LOW
                item['level'] = 1
            elif total <= medium_threshold:
                item['classification'] = 'SEDANG'
                item['color'] = COLOR_MEDIUM
                item['level'] = 2
            else:
                item['classification'] = 'TINGGI'
                item['color'] = COLOR_HIGH
                item['level'] = 3
        
        response_data = {
            'success': True,
            'data': {
                'areas': area_data,
                'thresholds': {
                    'low_max': int(low_threshold),
                    'medium_max': int(medium_threshold),
                    'high_min': int(medium_threshold) + 1
                },
                'statistics': {
                    'total_areas': len(area_data),
                    'low_count': sum(1 for a in area_data if a['classification'] == 'RENDAH'),
                    'medium_count': sum(1 for a in area_data if a['classification'] == 'SEDANG'),
                    'high_count': sum(1 for a in area_data if a['classification'] == 'TINGGI'),
                }
            }
        }
        
        return Response(response_data)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        
        return Response({
            'success': False,
            'message': 'Terjadi kesalahan server',
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# UPDATE fungsi area_statistics yang sudah ada
# Tambahkan klasifikasi di response

@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def area_statistics(request, area_id=None):
    """
    Endpoint untuk mendapatkan statistik kriminalitas berdasarkan area/desa
    UPDATED: Dengan klasifikasi tingkat kriminalitas
    """
    try:
        if not area_id:
            return Response({
                'success': False,
                'message': 'Area ID diperlukan'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Ambil data area
        try:
            area = Area.objects.select_related('desa', 'kecamatan').get(id=area_id)
        except Area.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Area tidak ditemukan'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Ambil desa_id dari area
        desa_id = area.desa_id
        
        # Query laporan kejahatan berdasarkan desa
        laporan_kejahatan = LaporanKejahatan.objects.filter(
            desa_id=desa_id
        ).select_related('jenis_kejahatan', 'nama_kejahatan', 'status')
        
        # Total kasus
        total_cases = laporan_kejahatan.count()
        
        # ============ TAMBAHAN: KLASIFIKASI AREA ============
        # Ambil semua total kasus untuk perhitungan threshold
        all_areas = Area.objects.all()
        all_case_counts = []
        
        for a in all_areas:
            count = LaporanKejahatan.objects.filter(desa_id=a.desa_id).count()
            all_case_counts.append(count)
        
        # Hitung threshold
        import numpy as np
        all_case_counts_sorted = sorted(all_case_counts)
        low_threshold = np.percentile(all_case_counts_sorted, 33) if all_case_counts else 0
        medium_threshold = np.percentile(all_case_counts_sorted, 67) if all_case_counts else 0
        
        # Klasifikasi area ini
        if total_cases <= low_threshold:
            classification = 'RENDAH'
            color = '#10b981'  # Hijau
            level = 1
        elif total_cases <= medium_threshold:
            classification = 'SEDANG'
            color = '#f59e0b'  # Kuning
            level = 2
        else:
            classification = 'TINGGI'
            color = '#ef4444'  # Merah
            level = 3
        # ====================================================
        
        # Kasus selesai dan proses
        solved = laporan_kejahatan.filter(status__nama__icontains='selesai').count()
        pending = total_cases - solved
        
        # Statistik berdasarkan jenis kejahatan
        by_type = laporan_kejahatan.values(
            'jenis_kejahatan__nama_jenis_kejahatan'
        ).annotate(
            count=Count('id')
        ).order_by('-count')[:5]
        
        by_type_data = [
            {
                'name': item['jenis_kejahatan__nama_jenis_kejahatan'],
                'value': item['count']
            }
            for item in by_type
        ]
        
        # Tren bulanan (12 bulan terakhir)
        twelve_months_ago = timezone.now() - timedelta(days=365)
        monthly_data = laporan_kejahatan.filter(
            tanggal_kejadian__gte=twelve_months_ago
        ).annotate(
            month=TruncMonth('tanggal_kejadian')
        ).values('month').annotate(
            count=Count('id')
        ).order_by('month')
        
        # Format data bulanan
        month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 
                      'Jul', 'Agu', 'Sep', 'Okt', 'Nov', 'Des']
        
        monthly_trend = []
        for item in monthly_data:
            month_num = item['month'].month
            monthly_trend.append({
                'month': month_names[month_num - 1],
                'cases': item['count']
            })
        
        # Kasus terbaru (top 10)
        recent_cases = laporan_kejahatan.order_by('-tanggal_kejadian')[:10]
        recent_cases_data = [
            {
                'id': case.id,
                'type': case.nama_kejahatan.nama,
                'location': case.alamat[:50] + '...' if len(case.alamat) > 50 else case.alamat,
                'date': case.tanggal_kejadian.isoformat(),
                'status': case.status.nama
            }
            for case in recent_cases
        ]
        
        # Statistik infrastruktur
        security_posts = PosKeamanan.objects.filter(desa_id=desa_id).count()
        cctvs = CCTV.objects.filter(desa_id=desa_id).count()
        
        # Format response - UPDATED dengan klasifikasi
        response_data = {
            'success': True,
            'data': {
                'area_info': {
                    'id': area.id,
                    'name': area.desa.nama if area.desa else area.wadmkd,
                    'kecamatan': area.kecamatan.nama if area.kecamatan else area.wadmkc,
                    'luas': float(area.luas) if area.luas else 0,
                    # TAMBAHAN: Klasifikasi
                    'classification': classification,
                    'classification_color': color,
                    'classification_level': level
                },
                'crime_stats': {
                    'totalCases': total_cases,
                    'solved': solved,
                    'pending': pending,
                    'byType': by_type_data,
                    'monthlyTrend': monthly_trend,
                    'recentCases': recent_cases_data
                },
                'infrastructure': {
                    'security_posts': security_posts,
                    'cctvs': cctvs
                },
                # TAMBAHAN: Info threshold
                'thresholds': {
                    'low_max': int(low_threshold),
                    'medium_max': int(medium_threshold),
                    'high_min': int(medium_threshold) + 1
                }
            }
        }
        
        return Response(response_data)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        
        return Response({
            'success': False,
            'message': 'Terjadi kesalahan server',
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def area_statistics(request, area_id=None):
    """
    Endpoint untuk mendapatkan statistik kriminalitas berdasarkan area/desa
    GET /api/map/area-statistics/<area_id>/
    
    Returns:
    {
        "success": true,
        "data": {
            "area_info": {
                "id": 1,
                "name": "Nama Desa",
                "kecamatan": "Nama Kecamatan",
                "luas": 10.5
            },
            "crime_stats": {
                "totalCases": 25,
                "solved": 15,
                "pending": 10,
                "byType": [
                    {"name": "Pencurian", "value": 10},
                    {"name": "Perampokan", "value": 8}
                ],
                "monthlyTrend": [
                    {"month": "Jan", "cases": 5},
                    {"month": "Feb", "cases": 8}
                ],
                "recentCases": [
                    {
                        "id": 1,
                        "type": "Pencurian",
                        "location": "Jl. Example",
                        "date": "2025-01-15",
                        "status": "Selesai"
                    }
                ]
            },
            "infrastructure": {
                "security_posts": 3,
                "cctvs": 5
            }
        }
    }
    """
    try:
        if not area_id:
            return Response({
                'success': False,
                'message': 'Area ID diperlukan'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Ambil data area
        try:
            area = Area.objects.select_related('desa', 'kecamatan').get(id=area_id)
        except Area.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Area tidak ditemukan'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Ambil desa_id dari area
        desa_id = area.desa_id
        
        # Query laporan kejahatan berdasarkan desa
        laporan_kejahatan = LaporanKejahatan.objects.filter(
            desa_id=desa_id
        ).select_related('jenis_kejahatan', 'nama_kejahatan', 'status')
        
        # Total kasus
        total_cases = laporan_kejahatan.count()
        
        # Kasus selesai dan proses
        solved = laporan_kejahatan.filter(status__nama__icontains='selesai').count()
        pending = total_cases - solved
        
        # Statistik berdasarkan jenis kejahatan
        by_type = laporan_kejahatan.values(
            'jenis_kejahatan__nama_jenis_kejahatan'
        ).annotate(
            count=Count('id')
        ).order_by('-count')[:5]
        
        by_type_data = [
            {
                'name': item['jenis_kejahatan__nama_jenis_kejahatan'],
                'value': item['count']
            }
            for item in by_type
        ]
        
        # Tren bulanan (12 bulan terakhir)
        from django.db.models.functions import TruncMonth
        from django.utils import timezone
        
        twelve_months_ago = timezone.now() - timedelta(days=365)
        monthly_data = laporan_kejahatan.filter(
            tanggal_kejadian__gte=twelve_months_ago
        ).annotate(
            month=TruncMonth('tanggal_kejadian')
        ).values('month').annotate(
            count=Count('id')
        ).order_by('month')
        
        # Format data bulanan
        month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 
                      'Jul', 'Agu', 'Sep', 'Okt', 'Nov', 'Des']
        
        monthly_trend = []
        for item in monthly_data:
            month_num = item['month'].month
            monthly_trend.append({
                'month': month_names[month_num - 1],
                'cases': item['count']
            })
        
        # Kasus terbaru (top 10)
        recent_cases = laporan_kejahatan.order_by('-tanggal_kejadian')[:10]
        recent_cases_data = [
            {
                'id': case.id,
                'type': case.nama_kejahatan.nama,
                'location': case.alamat[:50] + '...' if len(case.alamat) > 50 else case.alamat,
                'date': case.tanggal_kejadian.isoformat(),
                'status': case.status.nama
            }
            for case in recent_cases
        ]
        
        # Statistik infrastruktur
        security_posts = PosKeamanan.objects.filter(desa_id=desa_id).count()
        cctvs = CCTV.objects.filter(desa_id=desa_id).count()
        
        # Format response
        response_data = {
            'success': True,
            'data': {
                'area_info': {
                    'id': area.id,
                    'name': area.desa.nama if area.desa else area.wadmkd,
                    'kecamatan': area.kecamatan.nama if area.kecamatan else area.wadmkc,
                    'luas': float(area.luas) if area.luas else 0
                },
                'crime_stats': {
                    'totalCases': total_cases,
                    'solved': solved,
                    'pending': pending,
                    'byType': by_type_data,
                    'monthlyTrend': monthly_trend,
                    'recentCases': recent_cases_data
                },
                'infrastructure': {
                    'security_posts': security_posts,
                    'cctvs': cctvs
                }
            }
        }
        
        return Response(response_data)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        
        return Response({
            'success': False,
            'message': 'Terjadi kesalahan server',
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)